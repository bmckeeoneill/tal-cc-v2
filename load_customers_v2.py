"""
load_customers_v2.py — Load 6k NetSuite reference customers with embeddings.

Usage:
    cd /Users/brianoneill/Desktop/TAL_CC_clean
    source venv/bin/activate
    python3 load_customers_v2.py

- Reads latest .xlsx from Downloads matching *Customers*NAICS*.xlsx
- Filters SI/partner rows and duplicates
- Generates OpenAI embeddings (text-embedding-3-small, 1536-dim)
- Upserts into customers table by website URL
- Preserves existing UUIDs (locked customer relationships stay intact)
"""

import glob
import os
import re
import sys
import time
import uuid
import psycopg2
import psycopg2.extras
import openpyxl
import openai

# ── Secrets ───────────────────────────────────────────────────────────────────
secrets = open('.streamlit/secrets.toml').read()
DB_URL  = re.search(r'DATABASE_URL\s*=\s*"([^"]+)"', secrets).group(1)
OA_KEY  = re.search(r'\[openai\].*?api_key\s*=\s*"([^"]+)"', secrets, re.DOTALL).group(1)
oa      = openai.OpenAI(api_key=OA_KEY)

# ── Find file ─────────────────────────────────────────────────────────────────
SEARCH_PATHS = [
    os.path.expanduser("~/Downloads/*Customers*NAICS*.xlsx"),
    os.path.expanduser("~/Downloads/*Customers*NAICS*.xls.xlsx"),
    os.path.expanduser("~/Desktop/*Customers*NAICS*.xlsx"),
]
candidates = []
for pattern in SEARCH_PATHS:
    candidates.extend(glob.glob(pattern))

if not candidates:
    sys.exit("No customer file found. Place it in ~/Downloads and name it with 'Customers' and 'NAICS'.")

candidates = [c for c in candidates if not os.path.basename(c).startswith("~$")]
if not candidates:
    sys.exit("No customer file found (only temp files). Close Excel and retry.")
latest = sorted(candidates, key=os.path.getmtime)[-1]
print(f"Loading: {os.path.basename(latest)}")

# ── Parse Excel ───────────────────────────────────────────────────────────────
SI_PATTERNS = ["(SI)", "(GSI)", "(VAR)", "(ISV)", " Inc. (", " LLC (", " Ltd ("]

wb = openpyxl.load_workbook(latest, read_only=True)
ws = wb.active
next(ws.iter_rows(max_row=1))  # skip header row

rows = []
seen_websites = set()
seen_names    = set()
skipped_si    = 0
skipped_dupe  = 0
skipped_blank = 0

for row in ws.iter_rows(min_row=2, values_only=True):
    # col1 = customer name, col4 = SI name (ignore)
    name     = (row[1] or "").strip()
    website  = (row[2] or "").strip().rstrip("/").lower()
    naics    = (row[5] or "").strip()
    desc     = (row[6] or "").strip()   # References Descriptors
    ref_stat = (row[8] or "").strip()
    v_rank   = row[9] if isinstance(row[9], (int, float)) else None
    biz_type = (row[10] or "").strip()
    product  = (row[12] or "").strip()
    size     = (row[15] or "").strip()
    state    = (row[16] or "").strip()
    highlights = str(row[17] or "").strip()[:500]

    if not name:
        skipped_blank += 1
        continue

    # Filter SI/partner entries that leaked into customer column
    if any(p in name for p in SI_PATTERNS):
        skipped_si += 1
        continue

    # Deduplicate by website, then name
    dedup_key = website or name.lower()
    if dedup_key in seen_websites:
        skipped_dupe += 1
        continue
    seen_websites.add(dedup_key)

    naics_code = naics[:6].strip() if naics else ""
    naics_desc = naics[7:].strip() if len(naics) > 7 else ""

    rows.append({
        "company_name":           name,
        "website":                website or None,
        "naics_code":             naics_code or None,
        "naics_description":      naics_desc or None,
        "industry":               biz_type or None,
        "business_type":          biz_type or None,
        "state":                  state or None,
        "reference_status":       ref_stat or None,
        "v_rank":                 int(v_rank) if v_rank is not None else None,
        "references_descriptors": desc or None,
        "highlights":             highlights or None,
        "company_size":           size or None,
        "notes":                  highlights or None,
    })

print(f"  {len(rows)} usable rows ({skipped_si} SI filtered, {skipped_dupe} dupes, {skipped_blank} blank)")

# ── Generate embeddings ───────────────────────────────────────────────────────
def embed_batch(texts: list[str]) -> list[list[float]]:
    resp = oa.embeddings.create(model="text-embedding-3-small", input=texts)
    return [r.embedding for r in resp.data]

def make_embed_text(r: dict) -> str:
    parts = [
        r["company_name"],
        f"{r['naics_code']} {r['naics_description']}" if r.get("naics_code") else "",
        r.get("business_type") or "",
        (r.get("references_descriptors") or "")[:300],
        (r.get("highlights") or "")[:200],
    ]
    return " | ".join(p for p in parts if p)

print("Generating embeddings...")
BATCH = 100
embeddings = []
for i in range(0, len(rows), BATCH):
    batch_texts = [make_embed_text(r) for r in rows[i:i+BATCH]]
    embeddings.extend(embed_batch(batch_texts))
    print(f"  {min(i+BATCH, len(rows))}/{len(rows)}")
    time.sleep(0.1)

print(f"  {len(embeddings)} embeddings generated")

# ── Upsert to Supabase ────────────────────────────────────────────────────────
print("Upserting to Supabase...")
conn = psycopg2.connect(DB_URL)
cur  = conn.cursor()

# Load existing website → id map to preserve UUIDs
cur.execute("SELECT id, website, company_name FROM customers WHERE website IS NOT NULL")
existing_by_site = {row[1].lower().rstrip("/"): row[0] for row in cur.fetchall() if row[1]}
cur.execute("SELECT id, company_name FROM customers WHERE website IS NULL")
existing_by_name = {row[1].lower(): row[0] for row in cur.fetchall()}

inserted = 0
updated  = 0

for r, emb in zip(rows, embeddings):
    site = (r["website"] or "").lower().rstrip("/")
    name_lc = r["company_name"].lower()

    existing_id = existing_by_site.get(site) or existing_by_name.get(name_lc)

    emb_str = "[" + ",".join(str(x) for x in emb) + "]"

    if existing_id:
        cur.execute("""
            UPDATE customers SET
                company_name = %s, website = %s, naics_code = %s, naics_description = %s,
                industry = %s, business_type = %s, state = %s, reference_status = %s,
                v_rank = %s, references_descriptors = %s, highlights = %s,
                company_size = %s, notes = %s, embedding = %s::vector
            WHERE id = %s
        """, (
            r["company_name"], r["website"], r["naics_code"], r["naics_description"],
            r["industry"], r["business_type"], r["state"], r["reference_status"],
            r["v_rank"], r["references_descriptors"], r["highlights"],
            r["company_size"], r["notes"], emb_str, existing_id
        ))
        updated += 1
    else:
        new_id = str(uuid.uuid4())
        cur.execute("""
            INSERT INTO customers (
                id, company_name, website, naics_code, naics_description,
                industry, business_type, state, reference_status,
                v_rank, references_descriptors, highlights, company_size, notes, embedding
            ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s::vector)
        """, (
            new_id, r["company_name"], r["website"], r["naics_code"], r["naics_description"],
            r["industry"], r["business_type"], r["state"], r["reference_status"],
            r["v_rank"], r["references_descriptors"], r["highlights"],
            r["company_size"], r["notes"], emb_str
        ))
        inserted += 1

    if (inserted + updated) % 500 == 0:
        conn.commit()
        print(f"  {inserted + updated} processed...")

conn.commit()

# Final count
cur.execute("SELECT count(*) FROM customers")
total = cur.fetchone()[0]
conn.close()

print(f"\nDone.")
print(f"  Inserted: {inserted}")
print(f"  Updated:  {updated}")
print(f"  Total in DB: {total}")
